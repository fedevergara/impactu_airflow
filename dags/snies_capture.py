"""Capture raw SNIES institution records from HECAA into MongoDB."""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import requests
from airflow import DAG
from airflow.providers.mongo.hooks.mongo import MongoHook
from airflow.providers.standard.operators.python import PythonOperator

SNIES_IES_URL = "https://hecaa.mineducacion.gov.co/consultaspublicas/ies"
MONGO_DB = "snies_institutions"
COLLECTION_NAME = "institutions"


default_args = {
    "owner": "impactu",
    "depends_on_past": False,
    "start_date": datetime(2026, 1, 1),
    "email_on_failure": False,
    "email_on_retry": False,
    "retries": 2,
    "retry_delay": timedelta(minutes=10),
}


def _extract_viewstate(html: str) -> str:
    match = re.search(
        r'name="javax\.faces\.ViewState"[^>]*value="([^"]+)"',
        html,
    )
    if not match:
        raise RuntimeError("Could not find javax.faces.ViewState in HECAA page")
    return match.group(1)


def _download_institutions_excel(cache_dir: str, timeout: int = 60) -> str:
    log = logging.getLogger(__name__)
    Path(cache_dir).mkdir(parents=True, exist_ok=True)

    output_path = os.path.join(cache_dir, "Instituciones.xlsx")
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0"})

    page_response = session.get(SNIES_IES_URL, timeout=timeout)
    page_response.raise_for_status()
    viewstate = _extract_viewstate(page_response.text)

    payload = {
        "j_idt92": "j_idt92",
        "j_idt92:j_idt94": "",
        "javax.faces.ViewState": viewstate,
    }
    headers = {
        "Origin": "https://hecaa.mineducacion.gov.co",
        "Referer": SNIES_IES_URL,
        "Content-Type": "application/x-www-form-urlencoded",
    }

    download_response = session.post(
        SNIES_IES_URL,
        data=payload,
        headers=headers,
        timeout=timeout,
    )
    download_response.raise_for_status()

    content_disposition = download_response.headers.get("Content-Disposition", "")
    if "attachment" not in content_disposition.lower():
        content_type = download_response.headers.get("Content-Type", "")
        raise RuntimeError(
            "HECAA response did not look like a file download: "
            f"Content-Type={content_type!r}, Content-Disposition={content_disposition!r}"
        )

    content = download_response.content
    if not content.startswith(b"PK"):
        raise RuntimeError("Downloaded HECAA payload does not look like an .xlsx file")

    Path(output_path).write_bytes(content)
    log.info("Downloaded SNIES institutions Excel to %s (%d bytes)", output_path, len(content))
    return output_path


def _make_unique_headers(headers: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    unique_headers: list[str] = []

    for index, header in enumerate(headers, start=1):
        base = f"column_{index}" if header is None or str(header).strip() == "" else str(header)

        count = seen.get(base, 0)
        seen[base] = count + 1
        unique_headers.append(base if count == 0 else f"{base}_{count + 1}")

    return unique_headers


def _iter_excel_records(file_path: str, sheet_name: str | None = None) -> list[dict[str, Any]]:
    from openpyxl import load_workbook  # type: ignore[import-untyped]

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)

        try:
            headers = _make_unique_headers(list(next(rows)))
        except StopIteration:
            return []

        records: list[dict[str, Any]] = []
        for row in rows:
            if row is None or all(value is None for value in row):
                continue

            values = list(row)
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))

            records.append(dict(zip(headers, values[: len(headers)], strict=True)))

        return records
    finally:
        workbook.close()


def _insert_records(
    records: list[dict[str, Any]],
    mongo_conn_id: str,
    mongo_db: str,
    collection_name: str,
    replace_existing: bool,
    chunk_size: int,
) -> int:
    log = logging.getLogger(__name__)
    hook = MongoHook(mongo_conn_id=mongo_conn_id)
    client = hook.get_conn()
    collection = client[mongo_db][collection_name]

    if replace_existing:
        result = collection.delete_many({})
        log.info(
            "Deleted %d existing documents from %s.%s",
            result.deleted_count,
            mongo_db,
            collection_name,
        )

    inserted = 0
    for start in range(0, len(records), chunk_size):
        chunk = records[start : start + chunk_size]
        if not chunk:
            continue
        collection.insert_many(chunk, ordered=False)
        inserted += len(chunk)

    log.info(
        "Inserted %d raw SNIES institution records into %s.%s", inserted, mongo_db, collection_name
    )
    return inserted


def download_and_load_snies_institutions(
    mongo_conn_id: str = "mongodb_default",
    mongo_db: str = "snies_institutions",
    collection_name: str = "institutions",
    cache_dir: str | None = "/tmp/impactu_airflow_cache/snies",
    sheet_name: str | None = None,
    replace_existing: bool = True,
    chunk_size: int = 1000,
    **_kwargs: Any,
) -> dict[str, Any]:
    """Download the HECAA institutions spreadsheet and load each row as a raw MongoDB document."""
    airflow_home = os.environ.get("AIRFLOW_HOME", os.getcwd())
    cache_dir = cache_dir or os.path.join(airflow_home, "cache", "snies")

    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than zero")

    excel_path = _download_institutions_excel(cache_dir=cache_dir)
    records = _iter_excel_records(excel_path, sheet_name=sheet_name)
    inserted = _insert_records(
        records=records,
        mongo_conn_id=mongo_conn_id,
        mongo_db=mongo_db,
        collection_name=collection_name,
        replace_existing=replace_existing,
        chunk_size=chunk_size,
    )

    return {
        "excel_path": excel_path,
        "records": len(records),
        "inserted": inserted,
        "mongo_db": mongo_db,
        "collection_name": collection_name,
    }


with DAG(
    dag_id="snies_capture",
    default_args=default_args,
    description="Download raw SNIES institutions from HECAA and load into MongoDB",
    schedule="@monthly",
    catchup=False,
    is_paused_upon_creation=True,
    render_template_as_native_obj=True,
    tags=["capture", "snies", "institutions"],
) as dag:
    download_and_load_task = PythonOperator(
        task_id="download_snies_institutions",
        python_callable=download_and_load_snies_institutions,
        op_kwargs={
            "mongo_db": MONGO_DB,
            "collection_name": COLLECTION_NAME,
        },
    )
